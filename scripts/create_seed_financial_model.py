"""Create ChemBalance's assumption-led Seed financial model and chart.

This is a planning model for a pre-revenue company. It intentionally separates
management assumptions from linked calculations, uses no historical revenue,
and should be replaced with signed-customer and actual-expense data as available.
"""

from __future__ import annotations

from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension

ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = ROOT / "docs"
WORKBOOK_PATH = OUTPUT_DIR / "ChemBalance_Seed_Financial_Model.xlsx"
CHART_PATH = OUTPUT_DIR / "seed_financial_overview.png"
YEARS = ["2027E", "2028E", "2029E"]

# USD in thousands except explicitly designated prices.  All are planning
# assumptions created for this Seed model, not historical actuals or market data.
ASSUMPTIONS = {
    "New Pro subscriptions": [180, 600, 1400],
    "New Team accounts": [6, 20, 45],
    "New Enterprise accounts": [1, 4, 10],
    "Pro annual price ($)": [180, 180, 180],
    "Team annual contract value ($)": [1500, 1500, 1500],
    "Enterprise annual contract value ($)": [7500, 7500, 7500],
    "Cost of revenue (% of revenue)": [0.09, 0.09, 0.09],
    "Research & development ($000)": [210, 280, 350],
    "Sales & marketing ($000)": [90, 150, 240],
    "General & administrative ($000)": [70, 95, 130],
    "Working capital / tools ($000)": [20, 30, 40],
    "Seed capital proceeds ($000)": [1350, 0, 0],
}

# These unit economics are targets for scenario planning, not observed results.
UNIT_ECONOMICS = {
    "Pro": {"acv": 180, "gross_margin": 0.91, "annual_churn": 0.25, "cac": 90},
    "Team": {"acv": 1500, "gross_margin": 0.91, "annual_churn": 0.12, "cac": 900},
    "Enterprise": {"acv": 7500, "gross_margin": 0.91, "annual_churn": 0.08, "cac": 5000},
}

TITLE_FILL = PatternFill("solid", fgColor="135B44")
SECTION_FILL = PatternFill("solid", fgColor="CFE9E0")
GRAY_FILL = PatternFill("solid", fgColor="E7E5E4")
WHITE_FONT = Font(color="FFFFFF", bold=True, size=16)
TITLE_FONT = Font(color="FFFFFF", bold=True, size=16)
SUBTITLE_FONT = Font(bold=True, size=11)
INPUT_FONT = Font(color="0000FF")
FORMULA_FONT = Font(color="000000")
LINK_FONT = Font(color="008000")
HEADER_FONT = Font(bold=True)
ITALIC_FONT = Font(italic=True)
THIN = Side(style="thin", color="000000")
DOUBLE = Side(style="double", color="000000")
CURRENCY = '$#,##0.0;($#,##0.0);-'
PERCENT = '0.0%'
MULTIPLE = '0.0x'


def setup_sheet(ws, title: str, subtitle: str, unit: str) -> None:
    """Apply the finance-workbook layout skeleton and common print settings."""
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 20
    ws["C3"] = title
    ws["C3"].fill = TITLE_FILL
    ws["C3"].font = TITLE_FONT
    ws["C3"].alignment = Alignment(horizontal="left")
    ws.merge_cells("C3:F3")
    ws["C5"] = subtitle
    ws["C5"].font = SUBTITLE_FONT
    ws.merge_cells("C5:F5")
    ws["C6"] = unit
    ws["C6"].font = ITALIC_FONT
    ws.merge_cells("C6:F6")
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.oddFooter.center.text = f"{title} — Page &[Page]"


def auto_fit(ws) -> None:
    """Auto-size used columns, retaining only the required gutter widths."""
    for column in range(3, ws.max_column + 1):
        letter = get_column_letter(column)
        lengths = [len(str(cell.value or "")) for cell in ws[letter]]
        ws.column_dimensions[letter].width = min(max(max(lengths, default=10) + 2, 12), 46)


def add_section(ws, row: int, title: str, last_column: int = 6) -> None:
    ws.cell(row, 3, title)
    for column in range(3, last_column + 1):
        cell = ws.cell(row, column)
        cell.fill = SECTION_FILL
        cell.font = HEADER_FONT
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=last_column)


def add_input_comment(cell) -> None:
    cell.comment = Comment(
        "Source: Management scenario assumption, 2026-08-26, ChemBalance Seed planning model. "
        "This is not historical actuals, contracted revenue, or third-party market data.",
        "Manus AI",
    )


def add_year_headers(ws, row: int) -> None:
    ws.cell(row, 3, "Metric").font = HEADER_FONT
    for index, year in enumerate(YEARS, 4):
        cell = ws.cell(row, index, year)
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="right")


def make_assumptions(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Assumptions"
    setup_sheet(ws, "ChemBalance Seed Assumptions", "Management scenario inputs; replace blue cells with actual operating data.", "USD in $000 except prices and percentages")
    add_year_headers(ws, 8)
    row = 9
    for label, values in ASSUMPTIONS.items():
        ws.cell(row, 3, label)
        for position, value in enumerate(values, 4):
            cell = ws.cell(row, position, value)
            cell.font = INPUT_FONT
            cell.alignment = Alignment(horizontal="right")
            add_input_comment(cell)
            if "(%)" in label:
                cell.number_format = PERCENT
            elif "price" in label.lower() or "value" in label.lower():
                cell.number_format = '$#,##0;($#,##0);-'
            else:
                cell.number_format = CURRENCY if "$000" in label else '#,##0;(#,##0);-'
        row += 1
    add_section(ws, row + 1, "Unit economics targets", 7)
    headings = ["Segment", "ACV ($)", "Gross margin", "Annual churn", "CAC ($)", "LTV / CAC"]
    for column, heading in enumerate(headings, 3):
        ws.cell(row + 2, column, heading).font = HEADER_FONT
    for index, (segment, metrics) in enumerate(UNIT_ECONOMICS.items(), row + 3):
        ws.cell(index, 3, segment)
        for column, key in enumerate(("acv", "gross_margin", "annual_churn", "cac"), 4):
            cell = ws.cell(index, column, metrics[key])
            cell.font = INPUT_FONT
            cell.alignment = Alignment(horizontal="right")
            add_input_comment(cell)
            cell.number_format = PERCENT if key in {"gross_margin", "annual_churn"} else '$#,##0;($#,##0);-'
        ltv_cac = ws.cell(index, 8, f"=D{index}*E{index}/F{index}/G{index}")
        ltv_cac.font = FORMULA_FONT
        ltv_cac.number_format = MULTIPLE
    ws.print_area = f"B2:H{ws.max_row}"
    auto_fit(ws)


def make_revenue_build(wb: Workbook) -> None:
    ws = wb.create_sheet("Revenue Build")
    setup_sheet(ws, "ChemBalance Revenue Build", "Subscription revenue uses average active accounts for each forecast year.", "USD in $000 except account counts and annual contract values")
    add_year_headers(ws, 8)
    line_map = {
        "Pro subscriptions": (10, 11, 12, 13, 14, 15),
        "Team accounts": (19, 20, 21, 22, 23, 24),
        "Enterprise accounts": (28, 29, 30, 31, 32, 33),
    }
    for section, rows in line_map.items():
        add_section(ws, rows[0] - 1, section, 6)
        for label, row in zip(("Beginning accounts", "New accounts", "Ending accounts", "Average accounts", "Annual contract value ($)", "Revenue ($000)"), rows):
            ws.cell(row, 3, label)
    assumptions_rows = {"Pro subscriptions": (9, 12), "Team accounts": (10, 13), "Enterprise accounts": (11, 14)}
    for section, rows in line_map.items():
        begin, new, ending, average, price, revenue = rows
        new_assumption, price_assumption = assumptions_rows[section]
        for column in range(4, 7):
            year_offset = column - 4
            prior_column = get_column_letter(column - 1)
            current_column = get_column_letter(column)
            ws.cell(begin, column, "=0" if column == 4 else f"={prior_column}{ending}").font = FORMULA_FONT
            ws.cell(new, column, f"=Assumptions!{current_column}{new_assumption}").font = LINK_FONT
            ws.cell(ending, column, f"={current_column}{begin}+{current_column}{new}").font = FORMULA_FONT
            ws.cell(average, column, f"=({current_column}{begin}+{current_column}{ending})/2").font = FORMULA_FONT
            ws.cell(price, column, f"=Assumptions!{current_column}{price_assumption}").font = LINK_FONT
            ws.cell(revenue, column, f"={current_column}{average}*{current_column}{price}/1000").font = FORMULA_FONT
            for row in (begin, new, ending, average):
                ws.cell(row, column).number_format = '#,##0;(#,##0);-'
            ws.cell(price, column).number_format = '$#,##0;($#,##0);-'
            ws.cell(revenue, column).number_format = CURRENCY
    add_section(ws, 36, "Revenue summary", 6)
    ws.cell(37, 3, "Total revenue ($000)").font = Font(bold=True)
    for column in range(4, 7):
        letter = get_column_letter(column)
        cell = ws.cell(37, column, f"={letter}15+{letter}24+{letter}33")
        cell.font = FORMULA_FONT
        cell.number_format = CURRENCY
        cell.border = Border(top=THIN)
    ws.print_area = f"B2:F{ws.max_row}"
    auto_fit(ws)


def make_pnl_cash(wb: Workbook) -> None:
    ws = wb.create_sheet("P&L & Cash")
    setup_sheet(ws, "ChemBalance Profit & Loss and Cash", "Base-case cash plan funded by the Seed raise; no income tax benefit is modeled while loss-making.", "USD in $000")
    add_year_headers(ws, 8)
    rows = {
        "Revenue": 9,
        "Cost of revenue": 10,
        "Gross profit": 11,
        "Gross margin": 12,
        "Research & development": 14,
        "Sales & marketing": 15,
        "General & administrative": 16,
        "EBITDA": 17,
        "EBITDA margin": 18,
        "Working capital / tools": 20,
        "Free cash flow": 21,
        "Seed capital proceeds": 23,
        "Beginning cash": 24,
        "Ending cash": 25,
    }
    for label, row in rows.items():
        ws.cell(row, 3, label)
    for column in range(4, 7):
        letter = get_column_letter(column)
        previous = get_column_letter(column - 1)
        formulas = {
            rows["Revenue"]: f"='Revenue Build'!{letter}37",
            rows["Cost of revenue"]: f"=-{letter}{rows['Revenue']}*Assumptions!{letter}15",
            rows["Gross profit"]: f"={letter}{rows['Revenue']}+{letter}{rows['Cost of revenue']}",
            rows["Gross margin"]: f"={letter}{rows['Gross profit']}/{letter}{rows['Revenue']}",
            rows["Research & development"]: f"=-Assumptions!{letter}16",
            rows["Sales & marketing"]: f"=-Assumptions!{letter}17",
            rows["General & administrative"]: f"=-Assumptions!{letter}18",
            rows["EBITDA"]: f"=SUM({letter}{rows['Gross profit']}:{letter}{rows['General & administrative']})",
            rows["EBITDA margin"]: f"={letter}{rows['EBITDA']}/{letter}{rows['Revenue']}",
            rows["Working capital / tools"]: f"=-Assumptions!{letter}19",
            rows["Free cash flow"]: f"={letter}{rows['EBITDA']}+{letter}{rows['Working capital / tools']}",
            rows["Seed capital proceeds"]: f"=Assumptions!{letter}20",
            rows["Beginning cash"]: "=0" if column == 4 else f"={previous}{rows['Ending cash']}",
            rows["Ending cash"]: f"={letter}{rows['Beginning cash']}+{letter}{rows['Free cash flow']}+{letter}{rows['Seed capital proceeds']}",
        }
        for row, formula in formulas.items():
            cell = ws.cell(row, column, formula)
            cell.font = LINK_FONT if "!" in formula else FORMULA_FONT
            cell.alignment = Alignment(horizontal="right")
            cell.number_format = PERCENT if row in (rows["Gross margin"], rows["EBITDA margin"]) else CURRENCY
    for row in (rows["Revenue"], rows["Gross profit"], rows["EBITDA"], rows["Ending cash"]):
        for column in range(3, 7):
            ws.cell(row, column).font = Font(bold=True, color=ws.cell(row, column).font.color.rgb if ws.cell(row, column).font.color and ws.cell(row, column).font.color.type == 'rgb' else "000000")
            ws.cell(row, column).border = Border(top=THIN if row != rows["Ending cash"] else DOUBLE)
    for row in (rows["Gross margin"], rows["EBITDA margin"]):
        ws.cell(row, 3).font = ITALIC_FONT
    ws.print_area = f"B2:F{ws.max_row}"
    auto_fit(ws)


def make_scenarios(wb: Workbook) -> None:
    ws = wb.create_sheet("Scenarios")
    setup_sheet(ws, "ChemBalance Seed Scenario Summary", "Illustrative sensitivity of 2029 revenue and ending cash to new-logo acquisition pace.", "USD in $000 except percentages")
    headers = ["Scenario", "New-logo multiplier", "2029 revenue", "2029 ending cash", "Interpretation"]
    for index, label in enumerate(headers, 3):
        ws.cell(8, index, label).font = HEADER_FONT
    scenarios = [("Downside", 0.70, "Slower conversion or longer institutional sales cycle"), ("Base", 1.00, "Management operating plan"), ("Upside", 1.30, "Faster distribution or partner-assisted growth")]
    for row, (name, multiplier, interpretation) in enumerate(scenarios, 9):
        ws.cell(row, 3, name)
        mult = ws.cell(row, 4, multiplier)
        mult.font = INPUT_FONT
        mult.number_format = PERCENT
        add_input_comment(mult)
        ws.cell(row, 5, f"='Revenue Build'!F37*D{row}").font = LINK_FONT
        ws.cell(row, 5).number_format = CURRENCY
        # Expenses stay at base plan; this makes the sensitivity explicit rather than hiding cost changes.
        ws.cell(row, 6, f"='P&L & Cash'!F25+('Revenue Build'!F37*D{row}-'Revenue Build'!F37)*0.91").font = LINK_FONT
        ws.cell(row, 6).number_format = CURRENCY
        ws.cell(row, 7, interpretation)
        if name == "Base":
            for column in range(3, 8):
                ws.cell(row, column).fill = GRAY_FILL
    add_section(ws, 14, "Scenario interpretation", 7)
    ws["C15"] = "The downside / base / upside cases vary only customer acquisition. Product costs, headcount, and pricing remain at base assumptions so decision-makers can see commercial sensitivity clearly."
    ws.merge_cells("C15:G16")
    ws["C15"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.print_area = "B2:G16"
    auto_fit(ws)


def make_use_of_funds(wb: Workbook) -> None:
    ws = wb.create_sheet("Use of Funds")
    setup_sheet(ws, "ChemBalance Seed Use of Funds", "Proposed allocation of the Seed raise; management must align spending gates with demonstrated milestones.", "USD in $000")
    headers = ["Workstream", "% of Seed", "Amount", "Purpose and release gate"]
    for index, label in enumerate(headers, 3):
        ws.cell(8, index, label).font = HEADER_FONT
    uses = [
        ("Product & engineering", 0.45, "Stabilize multi-platform distribution; build percentage analysis and stoichiometry workspace."),
        ("Go-to-market & customer success", 0.25, "Run structured pilots, content-led acquisition, onboarding, and partner distribution experiments."),
        ("Operations & administration", 0.15, "Legal setup, finance, support operations, compliance and essential tools."),
        ("Security & distribution trust", 0.10, "Code-signing, macOS notarization, release integrity and production support."),
        ("Contingency reserve", 0.05, "Protect the plan against delayed conversion or platform certification work."),
    ]
    for row, (workstream, allocation, purpose) in enumerate(uses, 9):
        ws.cell(row, 3, workstream)
        percent = ws.cell(row, 4, allocation)
        percent.font = INPUT_FONT
        percent.number_format = PERCENT
        add_input_comment(percent)
        amount = ws.cell(row, 5, f"=Assumptions!D20*D{row}")
        amount.font = LINK_FONT
        amount.number_format = CURRENCY
        ws.cell(row, 6, purpose)
        ws.cell(row, 6).alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(15, 3, "Total Seed capital")
    ws.cell(15, 3).font = Font(bold=True)
    ws.cell(15, 4, "=SUM(D9:D13)")
    ws.cell(15, 4).font = FORMULA_FONT
    ws.cell(15, 4).number_format = PERCENT
    ws.cell(15, 5, "=SUM(E9:E13)")
    ws.cell(15, 5).font = FORMULA_FONT
    ws.cell(15, 5).number_format = CURRENCY
    for col in range(3, 6):
        ws.cell(15, col).border = Border(top=THIN, bottom=DOUBLE)
    ws.print_area = "B2:F15"
    auto_fit(ws)


def make_checks(wb: Workbook) -> None:
    ws = wb.create_sheet("Checks")
    setup_sheet(ws, "ChemBalance Seed Model Checks", "Formula checks; the model must be reviewed in Excel after all management inputs are updated.", "USD in $000 unless stated")
    checks = [
        ("Use of funds equals Seed proceeds", "='Use of Funds'!E15-Assumptions!D20", "Pass when 0.0"),
        ("Base 2029 ending cash", "='P&L & Cash'!F25", "Positive indicates cash coverage through 2029 under base assumptions"),
        ("2029 revenue growth vs. 2028", "='P&L & Cash'!F9/'P&L & Cash'!E9-1", "Review whether the implied sales ramp is supported by pilots and channel data"),
        ("2029 gross margin", "='P&L & Cash'!F12", "Derived from cost-of-revenue assumption"),
    ]
    ws["C8"] = "Check"
    ws["D8"] = "Value"
    ws["E8"] = "Interpretation"
    for cell in ws[8][2:5]:
        cell.font = HEADER_FONT
    for row, (label, formula, interpretation) in enumerate(checks, 9):
        ws.cell(row, 3, label)
        ws.cell(row, 4, formula).font = LINK_FONT
        ws.cell(row, 4).number_format = PERCENT if "growth" in label.lower() or "margin" in label.lower() else CURRENCY
        ws.cell(row, 5, interpretation)
        ws.cell(row, 5).alignment = Alignment(wrap_text=True)
    ws.print_area = "B2:E12"
    auto_fit(ws)


def create_chart() -> None:
    """Render an assumption-led, explicitly labeled revenue and cash overview."""
    pro_revenue = [16.2, 86.4, 266.4]
    team_revenue = [4.5, 24.0, 72.75]
    enterprise_revenue = [3.75, 22.5, 75.0]
    total_revenue = [24.45, 132.9, 414.15]
    ending_cash = [982.245, 548.189, 165.066]

    figure, (left, right) = plt.subplots(1, 2, figsize=(14, 6.5), dpi=180)
    figure.patch.set_facecolor("#F4F7FB")
    years = ["2027E", "2028E", "2029E"]
    colors = ["#0C8A66", "#467599", "#D90429"]
    bottom = [0, 0, 0]
    for values, label, color in zip((pro_revenue, team_revenue, enterprise_revenue), ("Pro", "Team", "Enterprise"), colors):
        left.bar(years, values, bottom=bottom, label=label, color=color)
        bottom = [a + b for a, b in zip(bottom, values)]
    left.set_title("Base-case recurring revenue build", fontweight="bold", color="#2F3E46")
    left.set_ylabel("Revenue ($000)", color="#2F3E46")
    left.legend(frameon=False)
    left.grid(axis="y", color="#DCE5E1")
    left.set_axisbelow(True)
    for index, value in enumerate(total_revenue):
        left.text(index, value + 10, f"${value:.1f}", ha="center", fontweight="bold", color="#2F3E46")

    right.plot(years, ending_cash, marker="o", color="#2F3E46", linewidth=3, markersize=9)
    right.fill_between(range(len(years)), ending_cash, color="#CFE9E0")
    right.set_xticks(range(len(years)), years)
    right.set_title("Base-case ending cash", fontweight="bold", color="#2F3E46")
    right.set_ylabel("Cash ($000)", color="#2F3E46")
    right.grid(axis="y", color="#DCE5E1")
    right.set_axisbelow(True)
    for index, value in enumerate(ending_cash):
        right.text(index, value + 35, f"${value:.1f}", ha="center", fontweight="bold", color="#2F3E46")

    for axis in (left, right):
        axis.set_facecolor("#FFFFFF")
        axis.spines[["top", "right"]].set_visible(False)
        axis.tick_params(colors="#2F3E46")
    figure.suptitle("ChemBalance Seed plan — management base case", fontsize=16, fontweight="bold", color="#2F3E46")
    figure.text(
        0.5,
        0.01,
        "Illustrative management assumptions only; not historical performance, contracted revenue, or an investment projection.",
        ha="center",
        fontsize=9,
        color="#59656B",
    )
    figure.tight_layout(rect=(0, 0.05, 1, 0.94))
    figure.savefig(CHART_PATH, bbox_inches="tight", facecolor=figure.get_facecolor())
    plt.close(figure)


def main() -> int:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    make_assumptions(workbook)
    make_revenue_build(workbook)
    make_pnl_cash(workbook)
    make_scenarios(workbook)
    make_use_of_funds(workbook)
    make_checks(workbook)
    workbook.save(WORKBOOK_PATH)
    create_chart()
    print(f"Created {WORKBOOK_PATH}")
    print(f"Created {CHART_PATH}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
