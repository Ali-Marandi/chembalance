"""Structural validation for the ChemBalance Seed planning workbook."""

from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
MODEL = ROOT / "docs" / "ChemBalance_Seed_Financial_Model.xlsx"
EXPECTED_SHEETS = ["Assumptions", "Revenue Build", "P&L & Cash", "Scenarios", "Use of Funds", "Checks"]


def main() -> int:
    workbook = load_workbook(MODEL, data_only=False)
    if workbook.sheetnames != EXPECTED_SHEETS:
        raise AssertionError(f"Unexpected worksheets: {workbook.sheetnames}")

    assumptions = workbook["Assumptions"]
    revenue = workbook["Revenue Build"]
    pnl = workbook["P&L & Cash"]
    scenarios = workbook["Scenarios"]
    funds = workbook["Use of Funds"]

    assert assumptions["D9"].comment is not None, "Management assumptions must carry an audit comment."
    assert revenue["D37"].value == "=D15+D24+D33", "Revenue total must link all segments."
    assert pnl["D9"].value == "='Revenue Build'!D37", "P&L revenue must link to revenue build."
    assert pnl["D25"].value == "=D24+D21+D23", "Ending cash must roll from cash flow and financing."
    assert funds["D15"].value == "=SUM(D9:D13)", "Use-of-funds allocation must sum by formula."
    assert scenarios["E10"].value == "='Revenue Build'!F37*D10", "Scenario revenue must link to base case."
    assert MODEL.stat().st_size > 10_000, "Workbook is unexpectedly small."
    print("Seed financial model structural checks passed.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
